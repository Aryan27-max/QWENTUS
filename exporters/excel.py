"""Excel report generation."""

from __future__ import annotations

from collections import Counter
from pathlib import Path
from statistics import mean
from datetime import datetime

from openpyxl import load_workbook

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import AtlasConfig
from models.candidate import ScreeningRecord


class ExcelExporter:
    """Create recruiter-friendly Excel workbooks."""

    HEADERS = [
        "Rank",
        "Candidate Name",
        "Email",
        "Phone",
        "Degree",
        "Overall Score",
        "GitHub Score",
        "Skills Score",
        "Projects Score",
        "Experience Score",
        "Achievements Score",
        "Leadership Score",
        "Communication Score",
        "Resume Quality Score",
        "Recommendation",
        "Decision",
        "Summary",
        "Resume Filename",
    ]

    def __init__(self, config: AtlasConfig) -> None:
        self.config = config

    def export(self, records: list[ScreeningRecord], stats=None) -> Path:
        """Write the screening report workbook to the Reports directory."""

        self.config.ensure_directories()
        report_path = self.config.paths.workbook
        workbook = self._build_workbook(records, stats)
        self._save_workbook(workbook, report_path)
        return report_path

    def load_existing_records(self) -> list[ScreeningRecord]:
        """Load the current workbook when resuming after a crash."""

        report_path = self.config.paths.workbook
        if not report_path.exists():
            return []

        try:
            workbook = load_workbook(report_path)
        except Exception:
            return []

        sheet = workbook["ALL CANDIDATES"] if "ALL CANDIDATES" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(min_row=1, values_only=True))
        workbook.close()
        if not rows:
            return []

        headers = [str(value) if value is not None else "" for value in rows[0]]
        records: list[ScreeningRecord] = []
        for row in rows[1:]:
            if not any(cell is not None and cell != "" for cell in row):
                continue
            values = {headers[index]: row[index] for index in range(min(len(headers), len(row)))}
            records.append(
                ScreeningRecord(
                    rank=int(values.get("Rank") or 0),
                    candidate_name=str(values.get("Candidate Name") or ""),
                    email=str(values.get("Email") or ""),
                    phone=str(values.get("Phone") or ""),
                    degree=str(values.get("Degree") or ""),
                    overall_score=int(values.get("Overall Score") or 0),
                    github_score=int(values.get("GitHub Score") or 0),
                    skills_score=int(values.get("Skills Score") or 0),
                    projects_score=int(values.get("Projects Score") or 0),
                    experience_score=int(values.get("Experience Score") or 0),
                    leadership_score=int(values.get("Leadership Score") or 0),
                    communication_score=int(values.get("Communication Score") or 0),
                    achievements_score=int(values.get("Achievements Score") or 0),
                    resume_quality_score=int(values.get("Resume Quality Score") or 0),
                    summary=str(values.get("Summary") or ""),
                    recommendation=str(values.get("Recommendation") or ""),
                    decision=str(values.get("Decision") or ""),
                    resume_file=str(values.get("Resume Filename") or values.get("Resume File") or ""),
                )
            )
        records.sort(key=lambda item: item.overall_score, reverse=True)
        for index, record in enumerate(records, start=1):
            record.rank = index
        return records

    def _build_workbook(self, records: list[ScreeningRecord], stats=None) -> Workbook:
        workbook = Workbook()
        workbook.remove(workbook.active)
        all_candidates = workbook.create_sheet("ALL CANDIDATES")
        shortlisted = workbook.create_sheet("SHORTLISTED")
        maybe = workbook.create_sheet("MAYBE")
        rejected = workbook.create_sheet("REJECTED")
        summary = workbook.create_sheet("SUMMARY")

        sorted_records = sorted(records, key=lambda item: item.overall_score, reverse=True)
        for index, record in enumerate(sorted_records, start=1):
            record.rank = index

        self._write_records(all_candidates, sorted_records)
        self._write_records(shortlisted, [record for record in sorted_records if record.decision == "Shortlisted"])
        self._write_records(maybe, [record for record in sorted_records if record.decision == "Maybe"])
        self._write_records(rejected, [record for record in sorted_records if record.decision == "Rejected"])
        self._write_summary(summary, sorted_records, stats)
        return workbook

    def _write_records(self, sheet, records: list[ScreeningRecord]) -> None:
        header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        for column_index, header in enumerate(self.HEADERS, start=1):
            cell = sheet.cell(row=1, column=column_index, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_index, record in enumerate(records, start=2):
            values = [
                record.rank,
                record.candidate_name,
                record.email,
                record.phone,
                record.degree,
                record.overall_score,
                record.github_score,
                record.skills_score,
                record.projects_score,
                record.experience_score,
                record.achievements_score,
                record.leadership_score,
                record.communication_score,
                record.resume_quality_score,
                record.recommendation,
                record.decision,
                record.summary,
                record.resume_file,
            ]
            for column_index, value in enumerate(values, start=1):
                sheet.cell(row=row_index, column=column_index, value=value)

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        self._autosize_columns(sheet)

    def _write_summary(self, sheet, records: list[ScreeningRecord], stats=None) -> None:
        total = len(records)
        scores = [record.overall_score for record in records]
        recommendations = Counter(record.recommendation for record in records)
        average_processing = round(getattr(stats, "average_processing_time", lambda: 0.0)(), 2) if stats else round(mean([record.processing_time_seconds for record in records]), 2) if records else 0
        average_llm = round(getattr(stats, "average_llm_time", lambda: 0.0)(), 2) if stats else 0
        average_ocr = round(getattr(stats, "average_ocr_time", lambda: 0.0)(), 2) if stats else 0
        average_scrape = round(getattr(stats, "average_scrape_time", lambda: 0.0)(), 2) if stats else 0
        average_excel = round(getattr(stats, "average_excel_time", lambda: 0.0)(), 2) if stats else 0
        total_runtime = round(getattr(stats, "total_runtime", lambda: 0.0)(), 2) if stats else 0

        rows = [
            ("Processed", total),
            ("Shortlisted", recommendations.get("Shortlisted", 0)),
            ("Maybe", recommendations.get("Maybe", 0)),
            ("Rejected", recommendations.get("Rejected", 0)),
            ("Failed", getattr(stats, "failed", 0) if stats else 0),
            ("Highest Score", max(scores) if scores else 0),
            ("Lowest Score", min(scores) if scores else 0),
            ("Average Score", round(mean(scores), 2) if scores else 0),
            ("Average Processing Time", average_processing),
            ("Average OCR Time", average_ocr),
            ("Average Scrape Time", average_scrape),
            ("Average LLM Time", average_llm),
            ("Average Excel Time", average_excel),
            ("Total Runtime", total_runtime),
            ("Timestamp", datetime.now().isoformat(sep=" ", timespec="seconds")),
        ]
        for row_index, (label, value) in enumerate(rows, start=1):
            sheet.cell(row=row_index, column=1, value=label)
            sheet.cell(row=row_index, column=2, value=value)

        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 18
        sheet.column_dimensions["C"].width = 18

    def _autosize_columns(self, sheet) -> None:
        for column_cells in sheet.columns:
            values = [len(str(cell.value)) for cell in column_cells if cell.value is not None]
            width = min(max(values, default=10) + 2, 45)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    def _save_workbook(self, workbook: Workbook, report_path: Path) -> None:
        for attempt in range(3):
            try:
                workbook.save(report_path)
                if report_path.exists():
                    load_workbook(report_path).close()
                    return
            except Exception:
                if attempt < 2:
                    workbook = self._build_workbook([], None)
                    continue
                raise
